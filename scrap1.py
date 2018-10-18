import selenium
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import time, openpyxl
from openpyxl import Workbook
import pymongo
from pymongo import MongoClient

#Getting the time for the creation
print (time.ctime())
driver = webdriver.Chrome()
#Saving data to a excelsheet
r=0
wb = Workbook()
ws = wb.create_sheet("Sheet1",0)
sheet = wb.active
#connect to database
client = MongoClient('localhost',27017)
#set the db object to pricetracker database
db = client['pricetracker']
#set the db to TV collection
collection = db['TV']
#result = collection.find()
for doc in collection.find({}):
	# changed variable names
	result = doc['url']
#Total value there in db
	xfile = openpyxl.load_workbook('price.xlsx')
	sheet = xfile.get_sheet_by_name('Sheet1')
	driver.get(result)
	if(result != 'null'):
		try:
			#Getting the title of the content
			title = driver.find_element_by_xpath('//*[@id="container"]/div/div[1]/div[2]/div/div[1]/div[2]/div[2]/div/div[1]/h1/span')
		except Exception:
			print ("No link to find title\t" + str(time.ctime()))
		try:
			for item in range(2,1,5):
				#Getting the price
				price = driver.find_element_by_xpath('//*[@id="container"]/div/div[1]/div[2]/div/div[1]/div[2]/div[2]/div/div['+item+']/div[1]/div/div[1]') or str('null')
		except TypeError:
			print ("Some error in finding the price\t" + str(time.ctime()))
		try:
			if(result>1):
				ws.cell(row = r, column = 0).value = title.text
				ws.cell(row = r, column = 1).value = price.text
				# worksheet.write(row,col,title.text)
				# worksheet.write(row,col+1,price.text)
				
				r += 1
			
				print ("increment") + str(r)
				wb.save('price.xlsx')

			else:
				print ("Some problem here at row: " +str(row) + str(time.ctime()))
		except Exception:
			pass
	print ("Data Scraped from url: ")
driver.quit()
print ("Completed") 
print (time.ctime())